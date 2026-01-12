"""Excel spreadsheet generation for financial data and models."""

import asyncio
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from bullsh.config import get_config
from bullsh.logging import log
from bullsh.tools.base import ToolResult, ToolStatus
from bullsh.tools.yahoo import scrape_yahoo

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _format_number(value: Any) -> str | float | None:
    """Format a number for display in Excel."""
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float)):
        return value
    return str(value)


def _format_large_number(value: Any) -> str:
    """Format large numbers with B/M suffix."""
    if value is None:
        return "N/A"
    if isinstance(value, str):
        return value
    if value >= 1_000_000_000_000:
        return f"${value / 1_000_000_000_000:.2f}T"
    if value >= 1_000_000_000:
        return f"${value / 1_000_000_000:.2f}B"
    if value >= 1_000_000:
        return f"${value / 1_000_000:.2f}M"
    return f"${value:,.0f}"


def _apply_header_style(ws, row: int, col_start: int, col_end: int) -> None:
    """Apply header styling to a row."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def _auto_column_width(ws) -> None:
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column = None
        for cell in column_cells:
            try:
                # Skip merged cells - they don't have column_letter
                if hasattr(cell, "column_letter"):
                    if column is None:
                        column = cell.column_letter
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        if column:
            ws.column_dimensions[column].width = min(max_length + 2, 50)


async def generate_excel(
    ticker: str,
    include_ratios: bool = True,
    compare_tickers: list[str] | None = None,
) -> ToolResult:
    """
    Generate Excel spreadsheet with financial data and models.

    Args:
        ticker: Primary stock ticker
        include_ratios: Include financial ratios sheet
        compare_tickers: Additional tickers for comparison

    Returns:
        ToolResult with path to generated Excel file
    """
    ticker = ticker.upper()
    all_tickers = [ticker]
    if compare_tickers:
        all_tickers.extend([t.upper() for t in compare_tickers])

    log("tools", f"generate_excel: Creating spreadsheet for {all_tickers}")

    try:
        # Fetch data for all tickers in parallel
        tasks = [scrape_yahoo(t) for t in all_tickers]
        results = await asyncio.gather(*tasks)

        # Check if we got any valid data
        ticker_data = {}
        for t, result in zip(all_tickers, results):
            if result.status == ToolStatus.SUCCESS:
                ticker_data[t] = result.data
            else:
                log("tools", f"generate_excel: Failed to get data for {t}: {result.error_message}")

        if not ticker_data:
            return ToolResult(
                data={},
                confidence=0.0,
                status=ToolStatus.FAILED,
                tool_name="generate_excel",
                ticker=ticker,
                error_message="Could not fetch data for any ticker",
            )

        # Create workbook
        wb = Workbook()

        # Sheet 1: Key Metrics
        ws_metrics = wb.active
        ws_metrics.title = "Key Metrics"
        _create_metrics_sheet(ws_metrics, ticker_data)

        # Sheet 2: Financial Ratios (if requested)
        if include_ratios:
            ws_ratios = wb.create_sheet("Financial Ratios")
            _create_ratios_sheet(ws_ratios, ticker_data)

        # Sheet 3: Comparison (if multiple tickers)
        if len(ticker_data) > 1:
            ws_compare = wb.create_sheet("Comparison")
            _create_comparison_sheet(ws_compare, ticker_data)

        # Sheet 4: Valuation Analysis
        ws_valuation = wb.create_sheet("Valuation")
        _create_valuation_sheet(ws_valuation, ticker_data)

        # Save workbook
        config = get_config()
        exports_dir = config.data_dir / "exports"
        exports_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if len(all_tickers) == 1:
            filename = f"{ticker}_financial_model_{timestamp}.xlsx"
        else:
            filename = f"comparison_{'_'.join(all_tickers)}_{timestamp}.xlsx"

        filepath = exports_dir / filename
        wb.save(filepath)

        log("tools", f"generate_excel: Saved to {filepath}")

        return ToolResult(
            data={
                "path": str(filepath),
                "filename": filename,
                "tickers": list(ticker_data.keys()),
                "sheets": [ws.title for ws in wb.worksheets],
            },
            confidence=1.0,
            status=ToolStatus.SUCCESS,
            tool_name="generate_excel",
            ticker=ticker,
        )

    except Exception as e:
        log("tools", f"generate_excel: Error - {e}", level="error")
        return ToolResult(
            data={},
            confidence=0.0,
            status=ToolStatus.FAILED,
            tool_name="generate_excel",
            ticker=ticker,
            error_message=str(e),
        )


def _create_metrics_sheet(ws, ticker_data: dict[str, dict]) -> None:
    """Create the Key Metrics sheet."""
    # Title
    ws["A1"] = "Key Metrics"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")

    # Headers
    row = 4
    metrics = [
        ("Ticker", "ticker"),
        ("Price", "price"),
        ("Previous Close", "previous_close"),
        ("Change %", "change_percent"),
        ("P/E Ratio", "pe_ratio"),
        ("Forward P/E", "forward_pe"),
        ("EPS", "eps"),
        ("Market Cap", "market_cap"),
        ("52W High", "52w_high"),
        ("52W Low", "52w_low"),
        ("Volume", "volume"),
        ("Avg Volume", "avg_volume"),
        ("Beta", "beta"),
        ("Dividend Yield", "dividend_yield"),
        ("Sector", "sector"),
        ("Industry", "industry"),
    ]

    # Write headers
    ws.cell(row=row, column=1, value="Metric")
    col = 2
    for ticker in ticker_data:
        ws.cell(row=row, column=col, value=ticker)
        col += 1
    _apply_header_style(ws, row, 1, len(ticker_data) + 1)

    # Write data
    for metric_name, metric_key in metrics:
        row += 1
        ws.cell(row=row, column=1, value=metric_name)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = BORDER

        col = 2
        for ticker, data in ticker_data.items():
            value = data.get(metric_key)
            cell = ws.cell(row=row, column=col)

            # Format specific metrics
            if metric_key == "market_cap":
                cell.value = _format_large_number(value)
            elif metric_key == "change_percent":
                if value is not None:
                    cell.value = f"{value:.2f}%"
                else:
                    cell.value = "N/A"
            elif metric_key == "dividend_yield":
                if value is not None:
                    # Yahoo returns dividend yield as decimal (0.025 = 2.5%)
                    cell.value = f"{value * 100:.2f}%"
                else:
                    cell.value = "N/A"
            elif metric_key in ("price", "previous_close", "52w_high", "52w_low", "eps"):
                if value is not None:
                    cell.value = f"${value:.2f}"
                else:
                    cell.value = "N/A"
            elif metric_key in ("pe_ratio", "forward_pe", "beta"):
                if value is not None:
                    cell.value = f"{value:.2f}"
                else:
                    cell.value = "N/A"
            elif metric_key in ("volume", "avg_volume"):
                if value is not None:
                    cell.value = f"{value:,.0f}"
                else:
                    cell.value = "N/A"
            else:
                cell.value = value if value is not None else "N/A"

            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right")
            col += 1

    _auto_column_width(ws)


def _create_ratios_sheet(ws, ticker_data: dict[str, dict]) -> None:
    """Create the Financial Ratios sheet with formulas."""
    ws["A1"] = "Financial Ratios"
    ws["A1"].font = Font(bold=True, size=14)

    row = 3
    ws.cell(row=row, column=1, value="Ratio")
    col = 2
    for ticker in ticker_data:
        ws.cell(row=row, column=col, value=ticker)
        col += 1
    _apply_header_style(ws, row, 1, len(ticker_data) + 1)

    ratios = [
        ("Valuation", None),  # Section header
        ("P/E Ratio (TTM)", "pe_ratio"),
        ("Forward P/E", "forward_pe"),
        ("Price to 52W High", None),  # Calculated
        ("Price to 52W Low", None),  # Calculated
        ("", None),  # Spacer
        ("Analyst Targets", None),  # Section header
        ("Target Mean Price", "target_mean_price"),
        ("Target High", "target_high_price"),
        ("Target Low", "target_low_price"),
        ("# of Analysts", "number_of_analysts"),
        ("Recommendation", "recommendation"),
    ]

    for ratio_name, ratio_key in ratios:
        row += 1

        # Section header
        if (
            ratio_key is None
            and ratio_name
            and ratio_name not in ("Price to 52W High", "Price to 52W Low")
        ):
            ws.cell(row=row, column=1, value=ratio_name)
            ws.cell(row=row, column=1).font = Font(bold=True, color="4472C4")
            continue

        ws.cell(row=row, column=1, value=ratio_name)
        ws.cell(row=row, column=1).border = BORDER

        col = 2
        for ticker, data in ticker_data.items():
            cell = ws.cell(row=row, column=col)

            # Calculate derived ratios
            if ratio_name == "Price to 52W High":
                price = data.get("price")
                high = data.get("52w_high")
                if price and high:
                    cell.value = f"{(price / high) * 100:.1f}%"
                else:
                    cell.value = "N/A"
            elif ratio_name == "Price to 52W Low":
                price = data.get("price")
                low = data.get("52w_low")
                if price and low:
                    cell.value = f"{(price / low) * 100:.1f}%"
                else:
                    cell.value = "N/A"
            elif ratio_key:
                value = data.get(ratio_key)
                if ratio_key in ("target_mean_price", "target_high_price", "target_low_price"):
                    cell.value = f"${value:.2f}" if value else "N/A"
                elif ratio_key in ("pe_ratio", "forward_pe"):
                    cell.value = f"{value:.2f}" if value else "N/A"
                else:
                    cell.value = value if value else "N/A"

            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right")
            col += 1

    _auto_column_width(ws)


def _create_comparison_sheet(ws, ticker_data: dict[str, dict]) -> None:
    """Create the Comparison sheet for multiple tickers."""
    ws["A1"] = "Side-by-Side Comparison"
    ws["A1"].font = Font(bold=True, size=14)

    tickers = list(ticker_data.keys())

    # Key comparison metrics
    comparison_metrics = [
        ("Price", "price", "${:.2f}"),
        ("Market Cap", "market_cap", "large"),
        ("P/E Ratio", "pe_ratio", "{:.2f}"),
        ("Forward P/E", "forward_pe", "{:.2f}"),
        ("EPS", "eps", "${:.2f}"),
        ("Beta", "beta", "{:.2f}"),
        ("Dividend Yield", "dividend_yield", "{:.2%}"),
        ("52W High", "52w_high", "${:.2f}"),
        ("52W Low", "52w_low", "${:.2f}"),
        ("Recommendation", "recommendation", "{}"),
    ]

    # Headers
    row = 3
    ws.cell(row=row, column=1, value="Metric")
    for i, ticker in enumerate(tickers):
        ws.cell(row=row, column=i + 2, value=ticker)
    ws.cell(row=row, column=len(tickers) + 2, value="Best")
    _apply_header_style(ws, row, 1, len(tickers) + 2)

    # Data rows
    for metric_name, metric_key, fmt in comparison_metrics:
        row += 1
        ws.cell(row=row, column=1, value=metric_name)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = BORDER

        values = []
        for i, ticker in enumerate(tickers):
            data = ticker_data[ticker]
            value = data.get(metric_key)
            cell = ws.cell(row=row, column=i + 2)

            if value is not None:
                if fmt == "large":
                    cell.value = _format_large_number(value)
                    values.append((value, ticker))
                elif "{:.2%}" in fmt:
                    cell.value = f"{value:.2%}" if value else "N/A"
                    values.append((value, ticker))
                elif "${" in fmt or "{:.2f}" in fmt:
                    cell.value = fmt.format(value)
                    values.append((value, ticker))
                else:
                    cell.value = str(value)
                    values.append((0, ticker))  # Non-numeric
            else:
                cell.value = "N/A"

            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right")

        # Determine "best" (highest for most metrics, lowest for P/E)
        best_cell = ws.cell(row=row, column=len(tickers) + 2)
        if values:
            if metric_key in ("pe_ratio", "forward_pe"):
                # Lower is better for P/E
                best = min(values, key=lambda x: x[0] if x[0] else float("inf"))
            elif metric_key == "recommendation":
                best_cell.value = "-"
            else:
                # Higher is better
                best = max(values, key=lambda x: x[0] if x[0] else float("-inf"))
            if metric_key != "recommendation":
                best_cell.value = best[1]
                best_cell.fill = POSITIVE_FILL
        best_cell.border = BORDER
        best_cell.alignment = Alignment(horizontal="center")

    _auto_column_width(ws)


def _create_valuation_sheet(ws, ticker_data: dict[str, dict]) -> None:
    """Create the Valuation Analysis sheet."""
    ws["A1"] = "Valuation Analysis"
    ws["A1"].font = Font(bold=True, size=14)

    row = 3
    for ticker, data in ticker_data.items():
        ws.cell(row=row, column=1, value=ticker)
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        # Price vs Targets
        price = data.get("price", 0)
        target_mean = data.get("target_mean_price")
        target_high = data.get("target_high_price")
        target_low = data.get("target_low_price")

        ws.cell(row=row, column=1, value="Current Price")
        ws.cell(row=row, column=2, value=f"${price:.2f}" if price else "N/A")
        row += 1

        if target_mean:
            upside = ((target_mean - price) / price * 100) if price else 0
            ws.cell(row=row, column=1, value="Target Mean")
            ws.cell(row=row, column=2, value=f"${target_mean:.2f}")
            ws.cell(row=row, column=3, value=f"{upside:+.1f}%")
            if upside > 0:
                ws.cell(row=row, column=3).fill = POSITIVE_FILL
            else:
                ws.cell(row=row, column=3).fill = NEGATIVE_FILL
            row += 1

        if target_high:
            upside = ((target_high - price) / price * 100) if price else 0
            ws.cell(row=row, column=1, value="Target High")
            ws.cell(row=row, column=2, value=f"${target_high:.2f}")
            ws.cell(row=row, column=3, value=f"{upside:+.1f}%")
            row += 1

        if target_low:
            downside = ((target_low - price) / price * 100) if price else 0
            ws.cell(row=row, column=1, value="Target Low")
            ws.cell(row=row, column=2, value=f"${target_low:.2f}")
            ws.cell(row=row, column=3, value=f"{downside:+.1f}%")
            row += 1

        # 52-week range analysis
        high_52w = data.get("52w_high")
        low_52w = data.get("52w_low")
        if high_52w and low_52w and price:
            range_position = (price - low_52w) / (high_52w - low_52w) * 100
            ws.cell(row=row, column=1, value="52W Range Position")
            ws.cell(row=row, column=2, value=f"{range_position:.1f}%")
            if range_position < 30:
                ws.cell(row=row, column=3, value="Near 52W Low")
                ws.cell(row=row, column=3).fill = POSITIVE_FILL
            elif range_position > 70:
                ws.cell(row=row, column=3, value="Near 52W High")
                ws.cell(row=row, column=3).fill = NEUTRAL_FILL
            row += 1

        row += 2  # Space between tickers

    _auto_column_width(ws)
