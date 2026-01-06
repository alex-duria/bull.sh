"""9-tab Excel workbook generation for factor analysis.

Uses openpyxl with styling from existing tools/excel.py patterns.
"""

import statistics
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from bullsh.config import get_config
from bullsh.factors.session import FactorState
from bullsh.factors.scenarios import SCENARIOS


# Styling constants (matching tools/excel.py)
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_factor_excel(
    state: FactorState,
    data: dict[str, Any],
    draft: bool = False,
) -> Path:
    """
    Generate 9-tab Excel workbook for factor analysis.

    Args:
        state: Factor session state with computed results
        data: Raw and computed data from fetching/calculation stages
        draft: If True, save as draft (in-progress analysis)

    Returns:
        Path to generated Excel file
    """
    ticker = state.primary_ticker or "UNKNOWN"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if draft:
        filename = f"{ticker}_factor_analysis_DRAFT_{timestamp}.xlsx"
    else:
        filename = f"{ticker}_factor_analysis_{timestamp}.xlsx"

    # Output directory
    config = get_config()
    output_dir = config.data_dir / "exports"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / filename

    # Create workbook
    wb = Workbook()

    # Create all 9 tabs
    _create_executive_summary(wb, state, data)
    _create_factor_exposures(wb, state, data)
    _create_peer_comparison(wb, state, data)
    _create_risk_decomposition(wb, state, data)
    _create_historical_exposures(wb, state, data)
    _create_scenario_analysis(wb, state, data)
    _create_fundamental_data(wb, state, data)
    _create_price_data(wb, state, data)
    _create_methodology(wb, state, data)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save
    wb.save(output_path)
    return output_path


def _style_header_row(ws, row: int, end_col: int) -> None:
    """Apply header styling to a row."""
    for col in range(1, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws) -> None:
    """Auto-adjust column widths."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _create_executive_summary(wb: Workbook, state: FactorState, data: dict[str, Any]) -> None:
    """Tab 1: Executive Summary."""
    ws = wb.create_sheet("Executive Summary", 0)

    ticker = state.primary_ticker or "UNKNOWN"
    yahoo_data = data.get("yahoo_data", {}).get(ticker, {})

    # Company info
    ws["A1"] = "Factor Analysis Report"
    ws["A1"].font = Font(bold=True, size=16)

    ws["A3"] = "Company Information"
    ws["A3"].font = Font(bold=True, size=12)

    info = [
        ("Ticker", ticker),
        ("Company", yahoo_data.get("company_name", ticker)),
        ("Sector", yahoo_data.get("sector", "N/A")),
        ("Industry", yahoo_data.get("industry", "N/A")),
        ("Market Cap", _format_market_cap(yahoo_data.get("market_cap"))),
        ("Price", f"${yahoo_data.get('price', 0):.2f}"),
        ("52W Range", f"${yahoo_data.get('52w_low', 0):.2f} - ${yahoo_data.get('52w_high', 0):.2f}"),
    ]

    for i, (label, value) in enumerate(info):
        ws.cell(row=4 + i, column=1, value=label)
        ws.cell(row=4 + i, column=2, value=value)

    # Factor profile table
    ws["A13"] = "Factor Profile"
    ws["A13"].font = Font(bold=True, size=12)

    headers = ["Factor", "Z-Score", "Percentile", "Interpretation"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=14, column=col, value=header)
    _style_header_row(ws, 14, 4)

    factor_scores = state.factor_scores.get(ticker, {})
    row = 15
    for factor in state.selected_factors:
        z_score = factor_scores.get(factor, 0)
        percentile = 50 + z_score * 15.87  # Approximate
        percentile = max(0, min(100, percentile))
        interpretation = _interpret_z_score(z_score)

        ws.cell(row=row, column=1, value=factor.title())
        ws.cell(row=row, column=2, value=round(z_score, 2))
        ws.cell(row=row, column=3, value=f"{percentile:.0f}%")
        ws.cell(row=row, column=4, value=interpretation)

        # Color coding
        if z_score > 0.5:
            ws.cell(row=row, column=2).fill = POSITIVE_FILL
        elif z_score < -0.5:
            ws.cell(row=row, column=2).fill = NEGATIVE_FILL

        row += 1

    # Bar chart for factor scores
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Factor Exposures"

    data_ref = Reference(ws, min_col=2, min_row=14, max_row=14 + len(state.selected_factors))
    cats_ref = Reference(ws, min_col=1, min_row=15, max_row=14 + len(state.selected_factors))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    chart.width = 12
    chart.height = 8

    ws.add_chart(chart, "F13")

    # Scenario summary
    ws["A25"] = "Scenario Analysis Summary"
    ws["A25"].font = Font(bold=True, size=12)

    headers = ["Scenario", "Expected Return"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=26, column=col, value=header)
    _style_header_row(ws, 26, 2)

    row = 27
    for scenario_name, expected_return in state.scenario_results.items():
        scenario = SCENARIOS.get(scenario_name)
        display_name = scenario.display_name if scenario else scenario_name
        ws.cell(row=row, column=1, value=display_name)
        ws.cell(row=row, column=2, value=f"{expected_return * 100:+.1f}%")

        if expected_return > 0:
            ws.cell(row=row, column=2).fill = POSITIVE_FILL
        elif expected_return < 0:
            ws.cell(row=row, column=2).fill = NEGATIVE_FILL

        row += 1

    _auto_width(ws)


def _create_factor_exposures(wb: Workbook, state: FactorState, data: dict[str, Any]) -> None:
    """Tab 2: Factor Exposures Detail."""
    ws = wb.create_sheet("Factor Exposures")

    ticker = state.primary_ticker or "UNKNOWN"
    ws["A1"] = f"Factor Exposures - {ticker}"
    ws["A1"].font = Font(bold=True, size=14)

    # Get profiles for component data
    profiles = data.get("profiles", {})
    primary_profile = profiles.get(ticker)

    row = 3
    for factor in state.selected_factors:
        ws.cell(row=row, column=1, value=factor.upper())
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        # Headers
        headers = ["Metric", "Value", "Peer Median", "Z-Score", "Percentile"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        _style_header_row(ws, row, 5)
        row += 1

        # Get factor score from profile if available
        factor_score = None
        if primary_profile and hasattr(primary_profile, 'scores'):
            factor_score = primary_profile.scores.get(factor)

        if factor_score:
            # Write component-level data
            for component_name, component_value in factor_score.components.items():
                if component_value is not None:
                    # Calculate peer median for this component
                    peer_values = []
                    for peer_ticker, peer_profile in profiles.items():
                        if peer_ticker != ticker and hasattr(peer_profile, 'scores'):
                            peer_score = peer_profile.scores.get(factor)
                            if peer_score and peer_score.components.get(component_name) is not None:
                                peer_values.append(peer_score.components[component_name])

                    peer_median = None
                    if peer_values:
                        peer_median = statistics.median(peer_values)

                    comp_z = factor_score.component_z_scores.get(component_name, 0)
                    comp_pct = 50 + comp_z * 15.87
                    comp_pct = max(0, min(100, comp_pct))

                    ws.cell(row=row, column=1, value=_format_metric_name(component_name))
                    ws.cell(row=row, column=2, value=round(component_value, 2))
                    ws.cell(row=row, column=3, value=round(peer_median, 2) if peer_median else "-")
                    ws.cell(row=row, column=4, value=round(comp_z, 2))
                    ws.cell(row=row, column=5, value=f"{comp_pct:.0f}%")

                    # Color-code z-score cell
                    if comp_z > 0.5:
                        ws.cell(row=row, column=4).fill = POSITIVE_FILL
                    elif comp_z < -0.5:
                        ws.cell(row=row, column=4).fill = NEGATIVE_FILL

                    row += 1

            # Composite row
            row += 1
            z_score = factor_score.z_score
            percentile = factor_score.percentile
        else:
            # Fallback to state-stored z-scores
            z_score = state.factor_scores.get(ticker, {}).get(factor, 0)
            percentile = 50 + z_score * 15.87
            percentile = max(0, min(100, percentile))

        ws.cell(row=row, column=1, value="Factor Composite")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=4, value=round(z_score, 2))
        ws.cell(row=row, column=5, value=f"{percentile:.0f}%")

        if z_score > 0.5:
            ws.cell(row=row, column=4).fill = POSITIVE_FILL
        elif z_score < -0.5:
            ws.cell(row=row, column=4).fill = NEGATIVE_FILL

        row += 2

        # Regression beta if available
        beta = state.regression_betas.get(ticker, {}).get(factor)
        if beta is not None:
            ws.cell(row=row, column=1, value="Regression Beta")
            ws.cell(row=row, column=2, value=round(beta, 3))
            row += 1

        row += 1

    _auto_width(ws)


def _format_metric_name(name: str) -> str:
    """Format component metric name for display."""
    replacements = {
        "pe_ratio": "P/E Ratio",
        "pb_ratio": "P/B Ratio",
        "ev_ebitda": "EV/EBITDA",
        "return_12m_1m": "12M Return (ex. last month)",
        "high_52w_proximity": "52-Week High Proximity",
        "roe": "Return on Equity",
        "debt_equity_inv": "Debt/Equity (inverted)",
        "earnings_stability": "Earnings Stability",
        "revenue_growth_yoy": "Revenue Growth YoY",
        "earnings_growth_yoy": "Earnings Growth YoY",
        "ln_market_cap": "Log Market Cap",
        "realized_vol_60d": "60-Day Volatility",
        "beta": "Beta",
    }
    return replacements.get(name, name.replace("_", " ").title())


def _create_peer_comparison(wb: Workbook, state: FactorState, data: dict[str, Any]) -> None:
    """Tab 3: Peer Comparison Matrix."""
    ws = wb.create_sheet("Peer Comparison")

    ws["A1"] = "Peer Comparison Matrix"
    ws["A1"].font = Font(bold=True, size=14)

    # Build comparison table
    all_tickers = [state.primary_ticker] + state.peers if state.primary_ticker else state.peers
    yahoo_data = data.get("yahoo_data", {})

    # Headers
    headers = ["Ticker", "Price", "Market Cap"] + [f.title() for f in state.selected_factors] + ["Composite", "Rank"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    _style_header_row(ws, 3, len(headers))

    # Data rows
    row = 4
    composite_scores = []

    for ticker in all_tickers:
        ticker_data = yahoo_data.get(ticker, {})
        factor_scores = state.factor_scores.get(ticker, {})

        # Calculate composite
        composite = sum(factor_scores.get(f, 0) * state.weights.get(f, 0) for f in state.selected_factors)
        composite_scores.append((ticker, composite))

        ws.cell(row=row, column=1, value=ticker)
        ws.cell(row=row, column=2, value=f"${ticker_data.get('price', 0):.2f}")
        ws.cell(row=row, column=3, value=_format_market_cap(ticker_data.get("market_cap")))

        col = 4
        for factor in state.selected_factors:
            z = factor_scores.get(factor, 0)
            ws.cell(row=row, column=col, value=round(z, 2))

            # Conditional formatting
            if z > 0.5:
                ws.cell(row=row, column=col).fill = POSITIVE_FILL
            elif z < -0.5:
                ws.cell(row=row, column=col).fill = NEGATIVE_FILL

            col += 1

        ws.cell(row=row, column=col, value=round(composite, 2))
        col += 1

        # Highlight primary ticker
        if ticker == state.primary_ticker:
            for c in range(1, col):
                ws.cell(row=row, column=c).font = Font(bold=True)

        row += 1

    # Add ranks
    composite_scores.sort(key=lambda x: x[1], reverse=True)
    rank_map = {t: i + 1 for i, (t, _) in enumerate(composite_scores)}

    for r in range(4, 4 + len(all_tickers)):
        ticker = ws.cell(row=r, column=1).value
        if ticker in rank_map:
            ws.cell(row=r, column=len(headers), value=rank_map[ticker])

    _auto_width(ws)


def _create_risk_decomposition(wb: Workbook, state: FactorState, data: dict[str, Any]) -> None:
    """Tab 4: Risk Decomposition."""
    ws = wb.create_sheet("Risk Decomposition")

    ws["A1"] = "Risk Decomposition"
    ws["A1"].font = Font(bold=True, size=14)

    # Variance decomposition table
    ws["A3"] = "Variance Decomposition"
    ws["A3"].font = Font(bold=True, size=12)

    headers = ["Factor", "Variance %"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    _style_header_row(ws, 4, 2)

    row = 5
    for factor, pct in state.variance_decomposition.items():
        ws.cell(row=row, column=1, value=factor.title())
        ws.cell(row=row, column=2, value=f"{pct:.1f}%")
        row += 1

    # Pie chart
    if state.variance_decomposition:
        chart = PieChart()
        chart.title = "Risk Attribution"

        data_ref = Reference(ws, min_col=2, min_row=4, max_row=4 + len(state.variance_decomposition))
        cats_ref = Reference(ws, min_col=1, min_row=5, max_row=4 + len(state.variance_decomposition))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 10
        chart.height = 8

        ws.add_chart(chart, "D3")

    # Correlation matrix
    ws["A20"] = "Correlation Matrix"
    ws["A20"].font = Font(bold=True, size=12)

    if state.correlation_matrix:
        tickers = list(state.correlation_matrix.keys())

        # Header row
        for col, ticker in enumerate(tickers, 2):
            ws.cell(row=21, column=col, value=ticker)
        _style_header_row(ws, 21, len(tickers) + 1)

        # Data rows
        for row_idx, ticker1 in enumerate(tickers):
            ws.cell(row=22 + row_idx, column=1, value=ticker1)
            for col_idx, ticker2 in enumerate(tickers):
                corr = state.correlation_matrix.get(ticker1, {}).get(ticker2, 0)
                ws.cell(row=22 + row_idx, column=2 + col_idx, value=round(corr, 2))

    _auto_width(ws)


def _create_historical_exposures(wb: Workbook, state: FactorState, data: dict[str, Any]) -> None:
    """Tab 5: Historical Factor Exposures (rolling charts)."""
    ws = wb.create_sheet("Historical Exposures")

    ticker = state.primary_ticker or "UNKNOWN"
    ws["A1"] = f"Historical Factor Exposures - {ticker}"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Rolling 36-Month Factor Betas"
    ws["A3"].font = Font(bold=True, size=12)

    rolling_betas = data.get("rolling_betas", {})

    if not rolling_betas:
        ws["A5"] = "Rolling factor exposures not available (requires 3+ years of data)"
        _auto_width(ws)
        return

    # Determine factor columns (exclude alpha for now)
    factor_names = [k for k in rolling_betas.keys() if k != "alpha"]

    if not factor_names:
        ws["A5"] = "No factor data available"
        _auto_width(ws)
        return

    # Get dates from first factor
    first_factor = factor_names[0]
    dates = [item[0] for item in rolling_betas[first_factor]]

    if not dates:
        ws["A5"] = "No rolling beta data available"
        _auto_width(ws)
        return

    # Write header row
    headers = ["Date"] + [f.replace("-", " ") for f in factor_names]
    for col, header in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=header)
    _style_header_row(ws, 5, len(headers))

    # Write data rows
    row = 6
    for i, date in enumerate(dates):
        ws.cell(row=row, column=1, value=date)
        for col, factor in enumerate(factor_names, 2):
            betas = rolling_betas.get(factor, [])
            if i < len(betas):
                ws.cell(row=row, column=col, value=round(betas[i][1], 3))
        row += 1

    # Create line chart for factor betas over time
    if len(dates) >= 2:
        chart = LineChart()
        chart.title = "Rolling Factor Betas (36-month window)"
        chart.style = 10
        chart.y_axis.title = "Beta"
        chart.x_axis.title = "Date"
        chart.width = 18
        chart.height = 10

        # Add data series for each factor
        data_end_row = 5 + len(dates)
        for col, factor in enumerate(factor_names, 2):
            data_ref = Reference(ws, min_col=col, min_row=5, max_row=data_end_row)
            chart.add_data(data_ref, titles_from_data=True)

        # Set categories (dates)
        cats = Reference(ws, min_col=1, min_row=6, max_row=data_end_row)
        chart.set_categories(cats)

        ws.add_chart(chart, f"{get_column_letter(len(headers) + 2)}3")

    # Summary statistics
    summary_row = row + 2
    ws.cell(row=summary_row, column=1, value="Summary Statistics")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)

    stat_headers = ["Factor", "Mean Beta", "Min Beta", "Max Beta", "Std Dev"]
    summary_row += 1
    for col, header in enumerate(stat_headers, 1):
        ws.cell(row=summary_row, column=col, value=header)
    _style_header_row(ws, summary_row, len(stat_headers))

    summary_row += 1
    for factor in factor_names:
        betas = [item[1] for item in rolling_betas.get(factor, [])]
        if betas:
            ws.cell(row=summary_row, column=1, value=factor.replace("-", " "))
            ws.cell(row=summary_row, column=2, value=round(statistics.mean(betas), 3))
            ws.cell(row=summary_row, column=3, value=round(min(betas), 3))
            ws.cell(row=summary_row, column=4, value=round(max(betas), 3))
            ws.cell(row=summary_row, column=5, value=round(statistics.stdev(betas), 3) if len(betas) > 1 else 0)
            summary_row += 1

    _auto_width(ws)


def _create_scenario_analysis(wb: Workbook, state: FactorState, data: dict[str, Any]) -> None:
    """Tab 6: Scenario Analysis."""
    ws = wb.create_sheet("Scenario Analysis")

    ticker = state.primary_ticker or "UNKNOWN"
    ws["A1"] = f"Scenario Analysis - {ticker}"
    ws["A1"].font = Font(bold=True, size=14)

    # Pre-built scenarios table
    ws["A3"] = "Scenario Expected Returns"
    ws["A3"].font = Font(bold=True, size=12)

    headers = ["Scenario", "Description", "Expected Return"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    _style_header_row(ws, 4, 3)

    row = 5
    for scenario_name, expected_return in state.scenario_results.items():
        scenario = SCENARIOS.get(scenario_name)
        if scenario:
            ws.cell(row=row, column=1, value=scenario.display_name)
            ws.cell(row=row, column=2, value=scenario.description[:80] + "...")
            ws.cell(row=row, column=3, value=f"{expected_return * 100:+.1f}%")

            if expected_return > 0:
                ws.cell(row=row, column=3).fill = POSITIVE_FILL
            elif expected_return < 0:
                ws.cell(row=row, column=3).fill = NEGATIVE_FILL

            row += 1

    # Factor assumptions table
    ws["A12"] = "Scenario Factor Assumptions"
    ws["A12"].font = Font(bold=True, size=12)

    headers = ["Scenario"] + [f.title() for f in state.selected_factors]
    for col, header in enumerate(headers, 1):
        ws.cell(row=13, column=col, value=header)
    _style_header_row(ws, 13, len(headers))

    row = 14
    for scenario_name, scenario in SCENARIOS.items():
        ws.cell(row=row, column=1, value=scenario.display_name)
        for col, factor in enumerate(state.selected_factors, 2):
            factor_return = scenario.factor_returns.get(factor, 0)
            ws.cell(row=row, column=col, value=f"{factor_return * 100:+.1f}%")
        row += 1

    # Custom scenario input section
    ws["A22"] = "Custom Scenario (Interactive)"
    ws["A22"].font = Font(bold=True, size=12)

    ws["A23"] = "Enter factor return assumptions below:"
    row = 24
    for factor in state.selected_factors:
        ws.cell(row=row, column=1, value=factor.title())
        ws.cell(row=row, column=2, value="0%")  # User editable
        row += 1

    _auto_width(ws)


def _create_fundamental_data(wb: Workbook, state: FactorState, data: dict[str, Any]) -> None:
    """Tab 7: Fundamental Data."""
    ws = wb.create_sheet("Fundamentals")

    ticker = state.primary_ticker or "UNKNOWN"
    ws["A1"] = f"Fundamental Data - {ticker}"
    ws["A1"].font = Font(bold=True, size=14)

    yahoo_data = data.get("yahoo_data", {}).get(ticker, {})

    # Key metrics
    metrics = [
        ("Market Cap", _format_market_cap(yahoo_data.get("market_cap"))),
        ("P/E Ratio", yahoo_data.get("pe_ratio")),
        ("Forward P/E", yahoo_data.get("forward_pe")),
        ("EPS", yahoo_data.get("eps")),
        ("Beta", yahoo_data.get("beta")),
        ("Dividend Yield", f"{(yahoo_data.get('dividend_yield') or 0) * 100:.2f}%"),
        ("52W High", f"${yahoo_data.get('52w_high', 0):.2f}"),
        ("52W Low", f"${yahoo_data.get('52w_low', 0):.2f}"),
        ("Average Volume", f"{yahoo_data.get('avg_volume', 0):,}"),
        ("Analyst Target", f"${yahoo_data.get('target_mean_price', 0):.2f}"),
        ("# Analysts", yahoo_data.get("number_of_analysts")),
        ("Recommendation", yahoo_data.get("recommendation")),
    ]

    ws["A3"] = "Key Metrics"
    ws["A3"].font = Font(bold=True, size=12)

    headers = ["Metric", "Value"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    _style_header_row(ws, 4, 2)

    for i, (label, value) in enumerate(metrics):
        ws.cell(row=5 + i, column=1, value=label)
        ws.cell(row=5 + i, column=2, value=value if value is not None else "N/A")

    _auto_width(ws)


def _create_price_data(wb: Workbook, state: FactorState, data: dict[str, Any]) -> None:
    """Tab 8: Price Data and Returns."""
    ws = wb.create_sheet("Price Data")

    ticker = state.primary_ticker or "UNKNOWN"
    ws["A1"] = f"Price Data - {ticker}"
    ws["A1"].font = Font(bold=True, size=14)

    price_history = data.get("price_history", {}).get(ticker, {})
    closes = price_history.get("closes", [])

    if not closes:
        ws["A3"] = "Price history not available"
        return

    # Monthly returns (limit to prevent huge files)
    returns = price_history.get("returns", [])
    monthly_returns = returns[::21][:36]  # ~monthly, last 3 years

    ws["A3"] = "Monthly Returns (Last 3 Years)"
    ws["A3"].font = Font(bold=True, size=12)

    headers = ["Date", "Return"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    _style_header_row(ws, 4, 2)

    for i, (date, ret) in enumerate(monthly_returns):
        ws.cell(row=5 + i, column=1, value=date)
        ws.cell(row=5 + i, column=2, value=f"{ret * 100:.2f}%")

        if ret > 0:
            ws.cell(row=5 + i, column=2).fill = POSITIVE_FILL
        elif ret < 0:
            ws.cell(row=5 + i, column=2).fill = NEGATIVE_FILL

    # Line chart for cumulative returns would go here

    _auto_width(ws)


def _create_methodology(wb: Workbook, state: FactorState, data: dict[str, Any]) -> None:
    """Tab 9: Methodology Notes."""
    ws = wb.create_sheet("Methodology")

    ws["A1"] = "Methodology Notes"
    ws["A1"].font = Font(bold=True, size=14)

    notes = [
        ("Factor Calculation", ""),
        ("", "Z-scores are calculated cross-sectionally within the peer group."),
        ("", "Z-score = (value - peer_mean) / peer_std"),
        ("", "Values are winsorized at 2.5th and 97.5th percentiles before calculation."),
        ("", ""),
        ("Factor Definitions", ""),
        ("Value", "Average z-score of inverted P/E, P/B, and EV/EBITDA. Higher = cheaper."),
        ("Momentum", "12-month return minus 1-month return, plus 52W high proximity."),
        ("Quality", "ROE, inverted debt/equity, earnings stability."),
        ("Growth", "YoY revenue growth and earnings growth."),
        ("Size", "Natural log of market capitalization."),
        ("Volatility", "60-day realized volatility and market beta."),
        ("", ""),
        ("Data Sources", ""),
        ("", "Price and fundamental data: Yahoo Finance (via yfinance)"),
        ("", "Factor returns: Ken French Data Library"),
        ("", ""),
        ("Glossary", ""),
        ("Z-score", "Standard deviations from the mean"),
        ("Factor exposure", "A stock's sensitivity to a systematic factor"),
        ("Idiosyncratic risk", "Stock-specific risk not explained by factors"),
        ("Cross-sectional", "Comparing across stocks at a point in time"),
        ("Winsorization", "Capping extreme values to reduce outlier impact"),
    ]

    row = 3
    for label, value in notes:
        if label:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=True) if not value else Font()
        if value:
            ws.cell(row=row, column=2, value=value)
        row += 1

    _auto_width(ws)


def _format_market_cap(market_cap: int | None) -> str:
    """Format market cap with appropriate suffix."""
    if market_cap is None:
        return "N/A"
    if market_cap >= 1e12:
        return f"${market_cap / 1e12:.2f}T"
    if market_cap >= 1e9:
        return f"${market_cap / 1e9:.2f}B"
    if market_cap >= 1e6:
        return f"${market_cap / 1e6:.2f}M"
    return f"${market_cap:,.0f}"


def _interpret_z_score(z: float) -> str:
    """Interpret a z-score value."""
    if z > 1.5:
        return "Very High"
    if z > 0.5:
        return "Above Average"
    if z > -0.5:
        return "Average"
    if z > -1.5:
        return "Below Average"
    return "Very Low"
